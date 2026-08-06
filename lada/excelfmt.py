"""Branded Excel composition.

Implements the guide-sheet house style:

* worksheet gridlines are switched **off** entirely, so column A carries no
  gridlines and no content - all rules are drawn deliberately;
* every section opens with a dark-neutral header block in **white bold** type,
  spanning the content columns, with the section title anchored in **column B**;
* section content starts in **column C**;
* interior table rules are light blue, and each section is enclosed by a
  dark-blue boundary;
* unformatted blank spacer rows separate sections;
* text is left-justified, top-aligned and word-wrapped, with estimated row
  heights so nothing is clipped.
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import config

BASE_FONT = "Segoe UI"
FIRST_CONTENT_COL = 3          # column C
TITLE_COL = 2                  # column B
SPACER_COL = 1                 # column A - always empty, never bordered


def _hex(colour: str) -> str:
    """openpyxl wants 'RRGGBB' or 'AARRGGBB' without the leading '#'."""
    return (config.normalise_hex(colour) or "#000000")[1:]


@dataclass
class SheetStyle:
    """Resolved colours + fonts for one workbook."""
    palette: dict[str, str]

    @property
    def dark(self) -> str:
        return self.palette["dark_neutral"]

    @property
    def blue(self) -> str:
        return self.palette["primary_blue"]

    @property
    def purple(self) -> str:
        return self.palette["primary_purple"]

    @property
    def teal(self) -> str:
        return self.palette["secondary_teal"]

    @property
    def light_blue(self) -> str:
        return self.palette.get("light_blue", config.lighten(self.blue, 0.70))

    @property
    def zebra(self) -> str:
        return self.palette.get("surface_alt", config.lighten(self.blue, 0.90))

    # --- fills ---
    def fill_dark(self) -> PatternFill:
        return PatternFill("solid", fgColor=_hex(self.dark))

    def fill_accent(self, colour: str | None = None) -> PatternFill:
        return PatternFill("solid", fgColor=_hex(colour or self.purple))

    def fill_zebra(self) -> PatternFill:
        return PatternFill("solid", fgColor=_hex(self.zebra))

    def fill_white(self) -> PatternFill:
        return PatternFill("solid", fgColor="FFFFFF")

    # --- fonts ---
    def font_section(self, size: int = 12) -> Font:
        return Font(name=BASE_FONT, size=size, bold=True, color="FFFFFF")

    def font_header(self, size: int = 10) -> Font:
        return Font(name=BASE_FONT, size=size, bold=True, color="FFFFFF")

    def font_body(self, size: int = 10, bold: bool = False,
                  colour: str | None = None) -> Font:
        return Font(name=BASE_FONT, size=size, bold=bold,
                    color=_hex(colour or self.dark))

    def font_muted(self, size: int = 9) -> Font:
        return Font(name=BASE_FONT, size=size, italic=True,
                    color=_hex(self.palette.get("text_muted", "#5A6B85")))

    # --- borders ---
    def grid(self) -> Border:
        side = Side(style="thin", color=_hex(self.light_blue))
        return Border(left=side, right=side, top=side, bottom=side)

    def grid_side(self) -> Side:
        return Side(style="thin", color=_hex(self.light_blue))

    def boundary_side(self) -> Side:
        return Side(style="medium", color=_hex(self.dark))


TOP_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)


class SectionSheet:
    """Cursor-based writer for one branded worksheet."""

    def __init__(self, worksheet: Worksheet, style: SheetStyle,
                 content_cols: int = 6, col_widths: Sequence[float] | None = None):
        self.ws = worksheet
        self.st = style
        self.content_cols = content_cols
        self.last_col = FIRST_CONTENT_COL + content_cols - 1
        self.row = 1
        self._section_no = 0

        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = _hex(style.purple)

        self.ws.column_dimensions[get_column_letter(SPACER_COL)].width = 2.6
        self.ws.column_dimensions[get_column_letter(TITLE_COL)].width = 3.4
        widths = list(col_widths or [])
        for offset in range(content_cols):
            letter = get_column_letter(FIRST_CONTENT_COL + offset)
            self.ws.column_dimensions[letter].width = (
                widths[offset] if offset < len(widths) else 26
            )

    # ------------------------------------------------------------ internals
    def _col_span_width(self, first: int, last: int) -> float:
        total = 0.0
        for col in range(first, last + 1):
            dim = self.ws.column_dimensions[get_column_letter(col)]
            total += float(dim.width or 8.43)
        return max(total, 8.0)

    def _estimate_height(self, text: str, first: int, last: int,
                         font_size: int = 10, min_lines: int = 1,
                         padding: float = 5.0) -> float:
        """Approximate wrapped row height in points (openpyxl cannot autofit)."""
        span = self._col_span_width(first, last)
        chars_per_line = max(8, int(span * (10.0 / max(font_size, 6)) * 1.02))
        lines = 0
        for para in str(text).split("\n"):
            lines += max(1, math.ceil(len(para) / chars_per_line))
        lines = max(lines, min_lines)
        return min(409.0, lines * (font_size * 1.42) + padding)

    def _outline(self, r1: int, c1: int, r2: int, c2: int) -> None:
        """Draw a dark-blue boundary around a block without touching column A."""
        side = self.st.boundary_side()
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                cell = self.ws.cell(row=row, column=col)
                border = cell.border
                cell.border = Border(
                    left=side if col == c1 else border.left,
                    right=side if col == c2 else border.right,
                    top=side if row == r1 else border.top,
                    bottom=side if row == r2 else border.bottom,
                )

    def _merge(self, row: int, first: int, last: int) -> None:
        if last > first:
            self.ws.merge_cells(start_row=row, start_column=first,
                                end_row=row, end_column=last)

    # -------------------------------------------------------------- public
    def spacer(self, count: int = 1, height: float = 8.0) -> None:
        """Insert blank, entirely unformatted rows as breathing space."""
        for _ in range(count):
            self.ws.row_dimensions[self.row].height = height
            self.row += 1

    def banner(self, title: str, subtitle: str = "",
               logo_path: Path | None = None) -> None:
        """Gradient-feel masthead block at the top of a sheet."""
        start = self.row
        self.ws.row_dimensions[self.row].height = 34
        cell = self.ws.cell(row=self.row, column=TITLE_COL, value=title)
        cell.font = Font(name=BASE_FONT, size=17, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self._merge(self.row, TITLE_COL, self.last_col)
        for col in range(TITLE_COL, self.last_col + 1):
            self.ws.cell(row=self.row, column=col).fill = self.st.fill_dark()
        self.row += 1

        if subtitle:
            self.ws.row_dimensions[self.row].height = 22
            cell = self.ws.cell(row=self.row, column=TITLE_COL, value=subtitle)
            cell.font = Font(name=BASE_FONT, size=10, bold=False, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            self._merge(self.row, TITLE_COL, self.last_col)
            for col in range(TITLE_COL, self.last_col + 1):
                self.ws.cell(row=self.row, column=col).fill = self.st.fill_accent(
                    self.st.purple)
            self.row += 1

        # Thin brand rule.
        self.ws.row_dimensions[self.row].height = 5
        for col in range(TITLE_COL, self.last_col + 1):
            self.ws.cell(row=self.row, column=col).fill = self.st.fill_accent(
                self.st.teal)
        self.row += 1
        self._outline(start, TITLE_COL, self.row - 1, self.last_col)

        if logo_path and Path(logo_path).exists():
            try:
                img = XLImage(str(logo_path))
                img.width = int(img.width * (46 / max(img.height, 1)))
                img.height = 46
                img.anchor = f"{get_column_letter(self._logo_anchor(img.width))}{start}"
                self.ws.add_image(img)
            except Exception:
                pass  # imagery is decorative; never fail the export for it
        self.spacer(1, 10)

    def _logo_anchor(self, pixel_width: int) -> int:
        """Right-most column the logo can start in and still fit inside the block.

        Anchoring by cell is all openpyxl offers, so walk the column widths from
        the right until there is room for the image; otherwise it lands on top of
        the banner title.
        """
        needed_chars = pixel_width / 7.0  # ~7 px per width unit at 11 pt
        accumulated = 0.0
        column = self.last_col
        for candidate in range(self.last_col, FIRST_CONTENT_COL, -1):
            accumulated += float(
                self.ws.column_dimensions[get_column_letter(candidate)].width or 8.43)
            column = candidate
            if accumulated >= needed_chars:
                break
        return column

    def meta_grid(self, pairs: Sequence[tuple[str, str]]) -> None:
        """Label/value metadata block: label in column C, value spanning the rest.

        One pair per row - packing two pairs across looked asymmetric because the
        sheet's column widths are tuned for the section tables, not for a
        symmetric label/value grid.
        """
        start = self.row
        value_first = FIRST_CONTENT_COL + 1
        for label, value in pairs:
            label_cell = self.ws.cell(row=self.row, column=FIRST_CONTENT_COL,
                                      value=str(label).upper())
            label_cell.font = Font(name=BASE_FONT, size=8, bold=True,
                                   color=_hex(self.st.purple))
            label_cell.alignment = LEFT_CENTER
            label_cell.fill = self.st.fill_zebra()
            label_cell.border = self.st.grid()

            value_cell = self.ws.cell(row=self.row, column=value_first,
                                      value=str(value))
            value_cell.font = self.st.font_body(10, bold=True)
            value_cell.alignment = LEFT_CENTER
            for col in range(value_first, self.last_col + 1):
                self.ws.cell(row=self.row, column=col).border = self.st.grid()
            self._merge(self.row, value_first, self.last_col)

            self.ws.row_dimensions[self.row].height = max(
                20.0, self._estimate_height(value, value_first, self.last_col, 10))
            self.row += 1
        self._outline(start, FIRST_CONTENT_COL, self.row - 1, self.last_col)
        self.spacer(1, 10)

    def section_header(self, number: int | str, title: str,
                       caption: str = "") -> int:
        """Dark-neutral section block; title text lives in column B."""
        start = self.row
        label = f"SECTION {number}" if str(number).strip() else ""
        self.ws.row_dimensions[self.row].height = 27

        title_cell = self.ws.cell(row=self.row, column=TITLE_COL,
                                  value=(f"{label}   |   {title}" if label else title))
        title_cell.font = self.st.font_section(12)
        title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                         indent=1)
        self._merge(self.row, TITLE_COL, self.last_col)
        for col in range(TITLE_COL, self.last_col + 1):
            self.ws.cell(row=self.row, column=col).fill = self.st.fill_dark()
        self.row += 1

        if caption:
            height = self._estimate_height(caption, FIRST_CONTENT_COL,
                                           self.last_col, 9, padding=6)
            self.ws.row_dimensions[self.row].height = height
            cap = self.ws.cell(row=self.row, column=FIRST_CONTENT_COL, value=caption)
            cap.font = self.st.font_muted(9)
            cap.alignment = TOP_LEFT_WRAP
            cap.fill = self.st.fill_zebra()
            self._merge(self.row, FIRST_CONTENT_COL, self.last_col)
            for col in range(FIRST_CONTENT_COL, self.last_col + 1):
                cell = self.ws.cell(row=self.row, column=col)
                cell.fill = self.st.fill_zebra()
                cell.border = self.st.grid()
            self.row += 1
        return start

    def paragraphs(self, text: str) -> None:
        """Prose content block starting in column C, word-wrapped."""
        blocks = [b.strip() for b in str(text or "").split("\n\n") if b.strip()]
        if not blocks:
            blocks = ["-"]
        for block in blocks:
            height = self._estimate_height(block, FIRST_CONTENT_COL,
                                           self.last_col, 10, padding=7)
            self.ws.row_dimensions[self.row].height = height
            cell = self.ws.cell(row=self.row, column=FIRST_CONTENT_COL, value=block)
            cell.font = self.st.font_body(10)
            cell.alignment = TOP_LEFT_WRAP
            self._merge(self.row, FIRST_CONTENT_COL, self.last_col)
            for col in range(FIRST_CONTENT_COL, self.last_col + 1):
                self.ws.cell(row=self.row, column=col).border = self.st.grid()
            self.row += 1

    def numbered_list(self, items: Iterable[str], *, prefix: str = "",
                      accent: str | None = None,
                      start_at: int = 1) -> None:
        """Numbered items: the number sits in column C, the text spans the rest."""
        for offset, item in enumerate(items, start=start_at):
            label = f"{prefix}{offset}"
            text = str(item).strip()
            height = self._estimate_height(text, FIRST_CONTENT_COL + 1,
                                           self.last_col, 10, padding=7)
            self.ws.row_dimensions[self.row].height = height

            num_cell = self.ws.cell(row=self.row, column=FIRST_CONTENT_COL,
                                    value=label)
            num_cell.font = Font(name=BASE_FONT, size=10, bold=True, color="FFFFFF")
            num_cell.alignment = Alignment(horizontal="center", vertical="top")
            num_cell.fill = self.st.fill_accent(accent or self.st.blue)
            num_cell.border = self.st.grid()

            body = self.ws.cell(row=self.row, column=FIRST_CONTENT_COL + 1,
                                value=text)
            body.font = self.st.font_body(10)
            body.alignment = TOP_LEFT_WRAP
            for col in range(FIRST_CONTENT_COL + 1, self.last_col + 1):
                self.ws.cell(row=self.row, column=col).border = self.st.grid()
            self._merge(self.row, FIRST_CONTENT_COL + 1, self.last_col)
            self.row += 1

    def sub_header(self, text: str, colour: str | None = None) -> None:
        """Secondary heading inside a section (e.g. 'Knowledge outcomes')."""
        self.ws.row_dimensions[self.row].height = 20
        cell = self.ws.cell(row=self.row, column=FIRST_CONTENT_COL, value=text)
        cell.font = Font(name=BASE_FONT, size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        self._merge(self.row, FIRST_CONTENT_COL, self.last_col)
        for col in range(FIRST_CONTENT_COL, self.last_col + 1):
            target = self.ws.cell(row=self.row, column=col)
            target.fill = self.st.fill_accent(colour or self.st.purple)
            target.border = self.st.grid()
        self.row += 1

    def table(self, headers: Sequence[str], rows: Sequence[Sequence[object]],
              *, widths: Sequence[float] | None = None,
              zebra: bool = True, total_row: Sequence[object] | None = None,
              align_center: Sequence[int] | None = None,
              span_last: bool = False) -> None:
        """Table starting in column C with a dark header and light-blue rules.

        ``span_last`` merges the final column through to the sheet's last content
        column, which lets a narrow table (e.g. 3 columns) fill the same block
        width as the wide tables on other sheets without fighting over the
        sheet-wide column widths.
        """
        if widths:
            for offset, width in enumerate(widths):
                col = FIRST_CONTENT_COL + offset
                if col <= self.last_col:
                    self.ws.column_dimensions[get_column_letter(col)].width = width

        centered = set(align_center or ())
        span = min(len(headers), self.content_cols)
        # Column index range each logical column occupies.
        def _bounds(offset: int) -> tuple[int, int]:
            first = FIRST_CONTENT_COL + offset
            if span_last and offset == span - 1:
                return first, self.last_col
            return first, first

        def _emit(values: Sequence[object], *, font: Font, fill: PatternFill | None,
                  height: float | None, centre_all: bool = False) -> None:
            heights = [18.0]
            for offset in range(span):
                first, last = _bounds(offset)
                value = values[offset] if offset < len(values) else ""
                cell = self.ws.cell(row=self.row, column=first,
                                    value=value if value is not None else "")
                cell.font = font
                cell.alignment = (CENTER_WRAP if (centre_all or offset in centered)
                                  else TOP_LEFT_WRAP)
                for col in range(first, last + 1):
                    target = self.ws.cell(row=self.row, column=col)
                    target.border = self.st.grid()
                    if fill is not None:
                        target.fill = fill
                self._merge(self.row, first, last)
                heights.append(self._estimate_height(value, first, last,
                                                     font.size or 10, padding=7))
            self.ws.row_dimensions[self.row].height = (
                height if height is not None else min(409.0, max(heights)))
            self.row += 1

        _emit(list(headers[:span]), font=self.st.font_header(9),
              fill=self.st.fill_dark(),
              height=max(24.0, max(
                  (self._estimate_height(h, *_bounds(i), 9, padding=8)
                   for i, h in enumerate(headers[:span])), default=24.0)),
              centre_all=True)

        for index, data_row in enumerate(rows):
            _emit(list(data_row), font=self.st.font_body(10),
                  fill=self.st.fill_zebra() if (zebra and index % 2 == 1) else None,
                  height=None)

        if total_row is not None:
            _emit(list(total_row),
                  font=Font(name=BASE_FONT, size=10, bold=True, color="FFFFFF"),
                  fill=self.st.fill_accent(self.st.teal), height=22.0)

    def close_section(self, start_row: int) -> None:
        """Enclose the section in a dark-blue boundary and add spacer rows."""
        self._outline(start_row, TITLE_COL, self.row - 1, self.last_col)
        self.spacer(2, 9)

    def freeze_below_banner(self, row: int | None = None) -> None:
        self.ws.freeze_panes = f"A{row or self.row}"

    def configure_print(self, footer_note: str = "") -> None:
        """Landscape, fit-to-width printing with a confidentiality footer.

        Without this the workbook is wider than a portrait page and every export
        (Excel's own 'Save as PDF' included) slices the tables across pages.
        """
        setup = self.ws.page_setup
        setup.orientation = "landscape"
        setup.paperSize = self.ws.PAPERSIZE_A4
        setup.fitToWidth = 1
        setup.fitToHeight = 0
        self.ws.sheet_properties.pageSetUpPr.fitToPage = True
        self.ws.print_options.horizontalCentered = True
        margins = self.ws.page_margins
        margins.left = margins.right = 0.3
        margins.top = margins.bottom = 0.45
        margins.header = margins.footer = 0.2
        self.ws.print_area = (
            f"A1:{get_column_letter(self.last_col)}{max(self.row - 1, 1)}"
        )
        self.ws.oddFooter.left.text = footer_note or config.CONFIDENTIALITY_NOTE
        self.ws.oddFooter.left.size = 7
        self.ws.oddFooter.left.color = _hex(self.st.palette.get("text_muted",
                                                                "#5A6B85"))
        self.ws.oddFooter.right.text = "Page &P of &N"
        self.ws.oddFooter.right.size = 7


def new_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def add_sheet(workbook: Workbook, title: str, style: SheetStyle,
              content_cols: int = 6,
              col_widths: Sequence[float] | None = None) -> SectionSheet:
    safe = title[:31].replace("/", "-").replace("\\", "-").replace("?", "")
    worksheet = workbook.create_sheet(safe)
    return SectionSheet(worksheet, style, content_cols, col_widths)


def save(workbook: Workbook, path: Path,
         sheets: Iterable[SectionSheet] = (),
         footer_note: str = "") -> Path:
    """Apply print setup to every sheet, then write the workbook."""
    for sheet in sheets:
        sheet.configure_print(footer_note)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
